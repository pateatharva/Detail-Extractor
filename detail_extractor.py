#Python program for "Multipurpose GUI: Phone Number Tracking, HTML Webscraping and Youtube Video Downloading"
#Group:
#Atharva Pate (58)
#Ajinkya Patil (59)
#Sreenandan Sivadas (71)
#Suraj Yeola (77)

from tkinter import *
import tkinter as tk
from customtkinter import *
import customtkinter as ctk
import webbrowser
import threading
import os
import subprocess
import mysql.connector
import pyttsx3
from PIL import Image
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

ctk.set_appearance_mode("dark")

root = ctk.CTk()

wb = Workbook()

#fonts
font1 = ctk.CTkFont("Press Start 2P", 30)
font2 = ctk.CTkFont("Pirata One", 30)
font3 = ctk.CTkFont("Gugi", 25)
font4 = ctk.CTkFont("Red Rose", 15)
font5 = ctk.CTkFont("Righteous", 40)
font6 = ctk.CTkFont("Righteous", 15)
font7 = ctk.CTkFont("Gugi", 15)
font8 = ctk.CTkFont("Red Rose", 20)

root.title("Detail Extractor")
root.geometry("600x400")
root.resizable(width = True, height = True)
root.iconbitmap('search.ico')

frame = ctk.CTkFrame(master = root)
frame.pack(pady = 10, padx = 10, fill = "both", side = "bottom", expand = True)

#frames
frame1 = ctk.CTkFrame(master = frame)                           #for phno tracker
frame1.pack(pady = 20, padx = 20, side = "left", fill = "both", expand = True)
frame2 = ctk.CTkFrame(master = frame)
frame2.pack(pady = 20, padx = 20, side = "right", fill = "both", expand = True)

#saving to MySQL
#global name
#con = mysql.connector.connect(host = "localhost", user = "root", password = "dan")
#cur = con.cursor(buffered = True)
#try:
#    cur.execute("use details")
#except:
#    cur.execute("create database details")
#    cur.execute("use details")
#try:
#    cur.execute("describe numbers")
#except:   
#    cur.execute("create table numbers(count int primary key auto_increment, name varchar(15), phno varchar(15), location varchar(15), timezone varchar(15), service_provider varchar(15), coordinates varchar(40))")

title = ctk.CTkLabel(root, text = "DETAIL EXTRACTOR", font = font1).pack(pady = 12, padx = 10)

#functions for phone number tracker
#def save():
#    cur.execute("insert into numbers(name, phno, location, timezone, service_provider, coordinates) values('{name}','{phoneno}','{location}','{time}','{service_name}','{lat} {lng}')")
#   con.commit()

def fileopen():
    print(os.getcwd())
    file = 'file:///' + os.getcwd() + '/' + 'location.html'
    webbrowser.open_new_tab(file)

def add_to_last_row(file_path,data):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    last_row = sheet.max_row + 1
    
    for col_num, value in enumerate(data, start=1):
        sheet.cell(row=last_row, column=col_num, value=value)

    wb.save(file_path)

def save():
    name = getname.get()
    data_to_add = [str(name), str(phoneno), str(location), str(time), str(service_name), str(lat) + "  " + str(lng)]
    file_path = 'Project1.xlsx'
    add_to_last_row(file_path, data_to_add)
    
def number():
    import phonenumbers 
    from phonenumbers import geocoder
    from phonenumbers import carrier
    from phonenumbers import timezone
    import sys
    from opencage.geocoder import OpenCageGeocode
    import folium
    from timezonefinder import TimezoneFinder
    global framephno
    global name
    global phoneno
    global location
    global time
    global service_name
    global lat
    global lng
    global data_to_add
    global getname
    global save_button
    
    top1 = ctk.CTkToplevel()
    top1.geometry("400x550")
    top1.title("Phone Number Tracker")

    framephno = ctk.CTkFrame(master = top1)
    framephno.pack(pady = 10, padx = 20, side = "bottom", fill = "both", expand = True)

    phnohead = ctk.CTkLabel(top1, text="Phone number Tracker", font = font2).pack(pady = 12, padx = 10)
    phoneno = e1.get()
    phone = ctk.CTkLabel(framephno, font = font4, text = "Phone number: " + phoneno).pack(pady = 5, padx = 10)

    key = "f650bbf2ee204f3f89afae9c734628f3" 
    phno = phonenumbers.parse(phoneno) 

    location = geocoder.description_for_number(phno, "en")
    located = ctk.CTkLabel(framephno, font = font4, text = "Location: " + location).pack(pady = 5, padx = 10) 

    time=timezone.time_zones_for_number(phno)
    timez=ctk.CTkLabel(framephno, font = font4, text = "Timezone: " + str(time)).pack(pady = 5, padx = 10)

    service_name=(carrier.name_for_number(phno,"en"))
    service = ctk.CTkLabel(framephno, font = font4, text = "Service provider: " + service_name).pack(pady = 5, padx = 10)

    geocoder = OpenCageGeocode(key)
    query = str(location)
    result = geocoder.geocode(query)
    lat = result[0]['geometry']['lat']
    lng = result[0]['geometry']['lng']
    coordinates = ctk.CTkLabel(framephno, font = font4, text = "Coordinates: " + str(lat) + " " + str(lng)).pack(pady = 5, padx = 10)

    def tts():
        engine = pyttsx3.init()
        voices = engine.getProperty('voices')
        engine.setProperty('voice', voices[1].id)
        engine.say("Welcome to  Phone Number Tracker. These are the details of the phone number," + phoneno + "This person resides in " + location + "with the timezone of " + str(time) + "and using the service provided by " + service_name + "The global coordinates of the location are " + str(lat) + "latitude and " + str(lng) + "longitude.    Have a good day")
        engine.runAndWait()

    tts_image = ctk.CTkImage(Image.open('white-speaker.png'))
    tts_button = ctk.CTkButton(framephno, corner_radius = 5, width = 0, fg_color = "black", image = tts_image, text = "", command = tts)
    tts_button.pack(pady = 12)

    my_map=folium.Map(location=[lat,lng],zoom_start=9)
    folium.Marker([lat,lng],popup=location).add_to(my_map)
    my_map.save("location.html")
    on_map = ctk.CTkLabel(framephno, font = font4, text_color = "green", text = "'location.html' was created").pack(pady = 12, padx = 10)
    ctk.CTkButton(framephno, text = "Open", border_width = 2, hover_color = "blue", border_color = "blue", fg_color = "#2b2b2b", command = fileopen).pack(pady = 12, padx = 10)

    getname = ctk.CTkEntry(framephno, width = 200, placeholder_text = "Name of target")
    getname.pack(pady = 12, padx = 10)
    save_button = ctk.CTkButton(framephno, border_width = 2, hover_color = "green", border_color = "green", fg_color = "#2b2b2b", text = "Save", command = save).pack(pady = 12,padx = 10)

#functions for webscraper
def parts():
    frame4 = ctk.CTkFrame(master = frame3)
    frame4.pack(pady = 10, padx = 10, side = "bottom", fill = "both", expand = True)
    
    text = ctk.CTkTextbox(frame4, font = font4, border_color = "green", border_width = 2, text_color = "green", width = 1000, height = 600)
    text.pack(pady = 10, padx = 10)
        
    choice = e21.get()
    
    if choice == 'whole':
        ans = soup.prettify()
        text.insert(END, ans)
    else:
        extracted=soup.find_all(choice)
        for extract in extracted:
            result = extract.text
            text.insert(END, result)

def open_downloaded_video(): 
    video_path = filedialog.askopenfilename()
    if video_path:
        if os.name == 'nt':
            os.startfile(video_path)
        else:
            subprocess.run(['open', video_path])

def utube_link(url):
    if utube_url.netloc == 'youtu.be':
        return True
    elif utube_url.netloc == 'youtube.com':
        return True
    else:
        return False

def download():
    from pytube import YouTube
    
    output_path = filedialog.askdirectory()

    if output_path:
        try:
            yt = YouTube(url, on_progress_callback = on_progress)
            video_stream = yt.streams.get_highest_resolution()
            video_stream.download(output_path)

            info_label.configure(text="Download complete!", text_color = "green")
            openvideo = ctk.CTkButton(frame3, text = "Open", hover_color = "green", fg_color = "#2b2b2b", border_width = 2, border_color = "green", command = open_downloaded_video)
            openvideo.pack(pady = 12, padx = 10)
            
        except Exception as e:
            info_label.configure(text=f"An error occurred: {str(e)}")
            
def on_progress(stream, chunk, bytes_remaining):
    total_size = stream.filesize
    bytes_downloaded = total_size - bytes_remaining
    percentage_of_completion = bytes_downloaded / total_size * 100
    per = str(int(percentage_of_completion))
    Percent.configure(text = per + '%')
    Percent.update()

    progressbar.set(float(percentage_of_completion) / 100)

def site():
    import requests
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse
    global e21
    global soup
    global choice
    global frame3
    global top2
    global frame5
    global utube_url
    global url
    global info_label
    global Percent
    global progressbar
    
    top2 = ctk.CTkToplevel()
    top2.title("Webscraper")
    
    frame3 = ctk.CTkFrame(master = top2)
    frame3.pack(pady = 20, padx = 20, side = "top", fill = "both", expand = True)
    frame5 = ctk.CTkFrame(master = frame3)
    frame5.pack(side = "bottom", pady = 20, padx = 20, fill = "both", expand = True)

    webs = ctk.CTkLabel(frame3, font = font5, text = "Webscraper").pack(side = "top", pady = 12, padx = 10)
    url = e2.get()
    link = ctk.CTkLabel(frame3, font = font4, text = "URL: " + url).pack(side = "top", pady = 12, padx = 10)
    response = requests.get(url)
    txt=response.text
    status = response.status_code
    status_is = ctk.CTkLabel(frame3, font = font4, text = "Status: " + str(status)).pack(side = "top", pady = 12, padx = 10)

    utube_url = urlparse(url)
    if utube_link(url):
        ctk.CTkLabel(frame3, text = "Download YouTube video", font = font8).pack(pady = 5, padx = 10)
        play = ctk.CTkImage(Image.open('play.ico'))
        download_location = ctk.CTkButton(frame3, text = "", corner_radius = 7, image = play, width = 80, border_width = 2, hover_color = "red", fg_color = "#2b2b2b", text_color = "white", border_color = "red", command = download)
        download_location.pack(pady = 12, padx = 10)

        Percent = ctk.CTkLabel(frame3, text = "0%")
        Percent.pack()
        progressbar = ctk.CTkProgressBar(frame3, width = 250, progress_color = "red")
        progressbar.set(0)
        progressbar.pack(padx = 10, pady = 10)
        
        info_label = ctk.CTkLabel(frame3, text = "")
        info_label.pack(pady = 12, padx = 10)
    else:
        pass
    
    soup=BeautifulSoup(txt,'html.parser')
    if status == 200:
        choose = ctk.CTkLabel(frame5, font = font4, text = "Enter desired html element\nOR\nEnter whole to extract complete\nhtml code").pack(side = "top", pady = 12, padx = 10)
        e21 = ctk.CTkEntry(frame5, placeholder_text = "'whole' or element")
        e21.pack(pady = 12, padx = 10)
        begin21 = ctk.CTkButton(frame5, hover_color = "grey", border_color = "black", fg_color = "white", text_color = "black", text = "Get Code", command = parts).pack(pady = 12, padx = 10)
         
    else:
        ctk.CTkLabel(frame5, font = font4, text = "Failed. Status code: " + str(status)).pack(side = "top", pady = 12, padx = 10)

#review website run
def run_streamlit():
    subprocess.run(["streamlit", "run", "review.py"])

def open_streamlit():
    streamlit_thread = threading.Thread(target=run_streamlit)
    streamlit_thread.start()

#functions for slide panel animation
class SlidePanel(ctk.CTkFrame):
    def __init__(self, parent, start_pos, end_pos):
        super().__init__(master = parent, fg_color = "black")
        self.start_pos = start_pos
        self.end_pos = end_pos
        self.width = abs(start_pos - end_pos)

        self.pos = self.start_pos
        self.in_start_pos = True
        
        self.place(relx = self.start_pos, rely = 0, relwidth = self.width, relheight = 1)

    def animate(self):
        if self.in_start_pos:
            self.animate_forward()
        else:
            self.animate_backward()

    def animate_forward(self):
        if self.pos > self.end_pos:
            self.pos -= 0.008
            self.place(relx = self.pos, rely = 0, relwidth = self.width, relheight = 1)
            self.after(6, self.animate_forward)
        else:
            self.in_start_pos = False

    def animate_backward(self):
         if self.pos < self.start_pos:
            self.pos += 0.008
            self.place(relx = self.pos, rely = 0, relwidth = self.width, relheight = 1)
            self.after(6, self.animate_backward)
         else:
            self.in_start_pos = True
            
#form for phone number tracker
head1 = ctk.CTkLabel(frame1, text = "Phone Number\n Tracker", font = font3).pack(pady = 28, padx = 10)
ctk.CTkLabel(frame1, font = font8, text = "Phone Number:").pack()

e1= ctk.CTkEntry(frame1, border_color = "grey", placeholder_text = "+001234567890", border_width = 2)
e1.pack(pady = 12, padx = 10)

begin1 = ctk.CTkButton(frame1, corner_radius = 6, text = "Locate", hover_color = "grey", border_width = 0 , border_color = "black", fg_color = "white", text_color = "black", command = number).pack(pady = 12, padx = 10)

#form for webscraper
head2 = ctk.CTkLabel(frame2, text = "Webscraper", font = font3).pack(pady = 40, padx = 20)
ctk.CTkLabel(frame2, font = font8, text = "URL:").pack()

e2= ctk.CTkEntry(frame2, border_color = "grey", placeholder_text = "https://abc.com", border_width = 2)
e2.pack(pady = 12, padx = 10)

begin2 = ctk.CTkButton(frame2, corner_radius = 6, text = "Search", hover_color = "grey", border_width = 0, border_color = "black", fg_color = "white", text_color = "black", command = site).pack(pady = 12, padx = 10)

#panel
animated_panel = SlidePanel(root, 0, -0.3)
panel_icon_img = ctk.CTkImage(Image.open('panel.png'))
panel_button = ctk.CTkButton(root, corner_radius = 2, width = 0, hover_color = "grey", fg_color = "black", image = panel_icon_img, text = "", command = animated_panel.animate)
panel_button.place(x = 0, y = 0)

ctk.CTkLabel(animated_panel, text = "This GUI and its\nbackend comprises\n100% python code.", font = font6, text_color = "grey").pack(expand = True, fill = 'both', pady = 12)
star = ctk.CTkImage(Image.open('star.ico'))
star_icon = ctk.CTkLabel(animated_panel, text = "", image = star).pack(pady = 2)
ctk.CTkLabel(animated_panel, text = "We would like to\nknow your views\n\nTo rate us,\nclick below.", font = font7).pack(expand = True, fill = 'both')
review = ctk.CTkButton(animated_panel, text = "R E V I E W", fg_color = "black", hover_color = "grey", text_color = "white", corner_radius = 0, font = font3, height = 10, command = open_streamlit).pack(expand = True, fill = 'both', pady = 5)
other_apps = ctk.CTkButton(animated_panel, text = "Try File Conversion app", fg_color = "black", text_color = "grey", hover_color = "black").pack(pady = 1)

root.mainloop()
