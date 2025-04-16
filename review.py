#Python program for review website

import streamlit as st
from streamlit_chat import message
from streamlit_option_menu import option_menu
from bardapi import Bard
import os
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import json
from streamlit_lottie import st_lottie
import requests

wb = Workbook()

os.environ["_BARD_API_KEY"] = "dQjxwoshL40QnglQtgvyxqVC_jht5NkWGI-Xy23tj8IKFfeOXMCh5jcjZfxDkGlNlbN9sg."

page_title = "Review our project"
page_icon = ":thumbsup:"

st.set_page_config(page_title = page_title, page_icon = page_icon)

def load_lottieurl(url:str):
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()

lottie = load_lottieurl("https://lottie.host/d8166163-c91b-4b2c-9924-047dc54d5019/26Ujf0Tihx.json")
lottie1 = load_lottieurl("https://lottie.host/7b7f9840-5566-4e24-b85a-3cdb4e0b63d4/knyEYe9sKx.json")

def submit_review(rating, comment):
    if comment == "":
        comment = "-----"
    if rating == 1:
        st.write(f"Rating: {rating}/5")
        st.write(f"Comment: {comment}")
    elif rating == 2:
        st.write(f"Rating: {rating}/5")
        st.write(f"Comment: {comment}")
    elif rating == 3:
        st.write(f"Rating: {rating}/5")
        st.write(f"Comment: {comment}")
    elif rating == 4:
        st.write(f"Rating: {rating}/5")
        st.write("Thank You")
        st.write(f"Comment: {comment}")
    elif rating == 5:
        st.write(f"Rating: {rating}/5")
        st.write("Thank You! You are the best")
        st.write(f"Comment: {comment}")
with st.sidebar:
    selected = option_menu(menu_title = "Directions", options = ["Review", "About Us"], icons = ["star-half", "person"], menu_icon = "map")

#save review data
def add_to_last_row(file_path,data):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        last_row = sheet.max_row + 1
        
        for col_num, value in enumerate(data, start=1):
            sheet.cell(row=last_row, column=col_num, value=value)

        wb.save(file_path)
            
def save():
    data_to_add = [str(name), str(rating), str(comment)]
    file_path = 'Review.xlsx'
    add_to_last_row(file_path, data_to_add)

if selected == "Review":
    st.title(f"{selected}")
    st_lottie(
        lottie, key = "star"
    )
    rating = st.slider("Rate out of 5", min_value=1, max_value=5, value=3, step=1)
    name = st.text_input("Pleaase enter your name")
    comment = st.text_area("Add a comment:")
    if st.button("Submit"):
        submit_review(rating, comment)
        save()
    
if selected == "About Us":
    st.title(f"{selected}")
    st_lottie(
        lottie1, key = "profile"
    )
    with st.container():
        st.subheader("Class: SE - B, Comp")
        st.title("Group Members: - ")
        st.write("Suraj Yeola(77)")
        st.write("Sreenandan Sivadas(71)")
        st.write("Ajinkya Patil(59)")
        st.write("Atharva Pate(58)")
        st.write("\nThank You")
